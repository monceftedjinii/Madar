from datetime import date, datetime, timedelta
from io import BytesIO
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from ..models import (
	Employee, Attendance, LeaveRequest, Task,
	AbsenceWarning, DisciplineFlag, Document, RoleChoices, AbsenceJustification
)


# ──────────────────────────────────────────────────────────────────────────────
# Date helpers
# ──────────────────────────────────────────────────────────────────────────────

EXPORT_LATE_THRESHOLD = datetime.strptime('09:00', '%H:%M').time()


def _parse_date_str(value):
	if not value:
		return None
	try:
		return date.fromisoformat(value)
	except ValueError:
		pass
	if '/' in value:
		parts = value.split('/')
		if len(parts) == 3:
			day, month, year = parts
			return date(int(year), int(month), int(day))
	raise ValueError('invalid date format')


def parse_date_range(request):
	"""Parse ?from= and ?to= query params. Default to current month."""
	from_param = request.query_params.get('from')
	to_param = request.query_params.get('to')
	today = date.today()

	from_date = _parse_date_str(from_param) if from_param else date(today.year, today.month, 1)

	if to_param:
		to_date = _parse_date_str(to_param)
	else:
		if today.month == 12:
			to_date = date(today.year + 1, 1, 1) - timedelta(days=1)
		else:
			to_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

	return from_date, to_date


def _format_time(value):
	return value.strftime('%H:%M') if value else '-'


def _format_date(value):
	return value.isoformat() if value else '-'


def _format_hours(value):
	return f"{value:.2f}"


def _parse_export_format(request):
	file_format = (
		request.query_params.get('file_format') or
		request.query_params.get('format') or
		'pdf'
	).lower()
	if file_format not in ('pdf', 'xlsx'):
		return None
	return file_format


# ──────────────────────────────────────────────────────────────────────────────
# Report-building helpers
# ──────────────────────────────────────────────────────────────────────────────

def _make_export_response(buffer, filename, content_type):
	response = HttpResponse(buffer.getvalue(), content_type=content_type)
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	return response


def _build_excel_report(title, headers, rows):
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = 'Report'

	worksheet.append([title])
	worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
	worksheet['A1'].font = Font(bold=True, size=14)
	worksheet['A1'].alignment = Alignment(horizontal='center')

	worksheet.append(headers)
	for cell in worksheet[2]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal='center')

	for row in rows:
		worksheet.append(row)

	from openpyxl.utils import get_column_letter
	for col_idx in range(1, len(headers) + 1):
		max_len = 0
		for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
			for value in cell:
				if value is not None:
					max_len = max(max_len, len(str(value)))
		worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

	output = BytesIO()
	workbook.save(output)
	output.seek(0)
	return output


def _build_pdf_report(title, headers, rows):
	output = BytesIO()
	doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=24, bottomMargin=24)
	styles = getSampleStyleSheet()
	story = [Paragraph(title, styles['Title']), Spacer(1, 12)]

	data = [headers] + rows
	table = Table(data, repeatRows=1)
	table.setStyle(TableStyle([
		('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
		('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
		('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
		('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
		('ALIGN', (0, 0), (-1, 0), 'CENTER'),
		('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
	]))
	story.append(table)
	doc.build(story)
	output.seek(0)
	return output


# ──────────────────────────────────────────────────────────────────────────────
# Scoping helpers
# ──────────────────────────────────────────────────────────────────────────────

def _resolve_employee_scope(user):
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return None
		return Employee.objects.filter(id=emp.id)
	if user.role == RoleChoices.CHEF:
		try:
			chef_emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return None
		return Employee.objects.filter(service_id=chef_emp.service_id)
	if user.role in [RoleChoices.GRH, RoleChoices.RH_SIMPLE, RoleChoices.RH_AGENT]:
		return Employee.objects.all()
	return None


def _get_dept_scope(user):
	"""Return the service_id for scoped roles (CHEF), None for global roles."""
	if user.role == RoleChoices.CHEF:
		try:
			emp = Employee.objects.get(email=user.email)
			return emp.service_id
		except Employee.DoesNotExist:
			return None
	if user.role in [RoleChoices.RH_SIMPLE, RoleChoices.GRH]:
		return None
	return None


# ──────────────────────────────────────────────────────────────────────────────
# Summary aggregation helpers
# ──────────────────────────────────────────────────────────────────────────────

def _count_employees(user):
	if user.role == RoleChoices.EMPLOYEE:
		return 1
	if user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		return Employee.objects.filter(service_id=service_id, is_active=True).count() if service_id else 0
	return Employee.objects.filter(is_active=True).count()


def _count_employees_in_period(user, from_date, to_date):
	"""Count distinct employees who had attendance records in the period."""
	qs = Attendance.objects.filter(date__gte=from_date, date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			return 1 if qs.filter(employee_id=emp.id).exists() else 0
		except Employee.DoesNotExist:
			return 0
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	else:
		try:
			emp = Employee.objects.get(email=user.email)
			if emp.service_id:
				qs = qs.filter(employee__service_id=emp.service_id)
		except Employee.DoesNotExist:
			pass
	return qs.values('employee_id').distinct().count()


def _count_attendance(user, from_date, to_date):
	qs = Attendance.objects.filter(date__gte=from_date, date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			return 0
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	return qs.count()


def _count_warnings(user, from_date, to_date):
	qs = AbsenceWarning.objects.filter(date__gte=from_date, date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			return 0
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	return qs.count()


def _count_absences(user, from_date, to_date):
	"""Count absence days (AbsenceJustification records) in the period."""
	qs = AbsenceJustification.objects.filter(date__gte=from_date, date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			return 0
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	else:
		# GRH/DRH — filter by their service (HR)
		try:
			emp = Employee.objects.get(email=user.email)
			if emp.service_id:
				qs = qs.filter(employee__service_id=emp.service_id)
		except Employee.DoesNotExist:
			pass
	return qs.count()


def _count_discipline_flags(user, from_date, to_date):
	qs = DisciplineFlag.objects.filter(month__gte=from_date, month__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			return 0
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	return qs.count()


def _count_leaves(user, from_date, to_date):
	qs = LeaveRequest.objects.filter(start_date__gte=from_date, start_date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			qs = qs.none()
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(employee__service_id=service_id) if service_id else qs.none()
	return {
		'leaves_pending_count': qs.filter(status=LeaveRequest.Status.PENDING).count(),
		'leaves_accepted_count': qs.filter(status=LeaveRequest.Status.ACCEPTED).count(),
		'leaves_refused_count': qs.filter(status=LeaveRequest.Status.REFUSED).count(),
	}


def _count_documents(user, from_date, to_date):
	qs = Document.objects.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
	if user.role == RoleChoices.EMPLOYEE:
		qs = qs.filter(created_by=user)
	elif user.role == RoleChoices.CHEF:
		service_id = _get_dept_scope(user)
		qs = qs.filter(source_service_id=service_id) if service_id else qs.none()
	return {
		'documents_created_count': qs.count(),
		'documents_validated_count': qs.filter(status=Document.Status.VALIDATED).count(),
	}


# ──────────────────────────────────────────────────────────────────────────────
# API Endpoints
# ──────────────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_summary(request):
	"""
	GET /api/reports/summary/?from=YYYY-MM-DD&to=YYYY-MM-DD
	Returns aggregated KPI counts scoped by user role and department.
	"""
	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	data = {
		'employees_count': _count_employees_in_period(request.user, from_date, to_date),
		'attendance_days_count': _count_attendance(request.user, from_date, to_date),
		'absences_detected_count': _count_absences(request.user, from_date, to_date),
		'warnings_count': _count_warnings(request.user, from_date, to_date),
		'discipline_flags_count': _count_discipline_flags(request.user, from_date, to_date),
	}
	data.update(_count_leaves(request.user, from_date, to_date))
	data.update(_count_documents(request.user, from_date, to_date))
	data['from'] = from_date.isoformat()
	data['to'] = to_date.isoformat()
	data['user_role'] = request.user.role
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_report(request):
	"""Export attendance records as PDF or Excel."""
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	employees_qs = _resolve_employee_scope(request.user)
	if employees_qs is None:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	employee_id = request.query_params.get('employee_id')
	if employee_id:
		if request.user.role not in (RoleChoices.CHEF, RoleChoices.GRH):
			return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)
		employees_qs = employees_qs.filter(id=employee_id)

	attendance_qs = Attendance.objects.filter(
		employee__in=employees_qs,
		date__gte=from_date,
		date__lte=to_date,
	).select_related('employee')

	attendance_map = {(a.employee_id, a.date): a for a in attendance_qs}
	employees = list(employees_qs.order_by('last_name', 'first_name'))

	rows = []
	current_date = from_date
	while current_date <= to_date:
		for emp in employees:
			attendance = attendance_map.get((emp.id, current_date))
			if attendance and attendance.check_in_time:
				status_label = 'Late' if attendance.check_in_time > EXPORT_LATE_THRESHOLD else 'Present'
			else:
				status_label = 'Absent'

			worked_hours = 0.0
			if attendance and attendance.check_in_time and attendance.check_out_time:
				start_dt = datetime.combine(current_date, attendance.check_in_time)
				end_dt = datetime.combine(current_date, attendance.check_out_time)
				if end_dt < start_dt:
					end_dt += timedelta(days=1)
				worked_hours = round((end_dt - start_dt).total_seconds() / 3600, 2)

			rows.append([
				f"{emp.first_name} {emp.last_name}",
				_format_date(current_date),
				_format_time(attendance.check_in_time if attendance else None),
				_format_time(attendance.check_out_time if attendance else None),
				_format_hours(worked_hours),
				status_label,
			])
		current_date += timedelta(days=1)

	headers = ['Employee', 'Date', 'Check-in', 'Check-out', 'Worked Hours', 'Status']
	title = f"Attendance Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"attendance_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		return _make_export_response(
			_build_excel_report(title, headers, rows),
			filename,
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
	return _make_export_response(_build_pdf_report(title, headers, rows), filename, 'application/pdf')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leaves_report(request):
	"""Export leave requests as PDF or Excel."""
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	user = request.user
	qs = LeaveRequest.objects.all().select_related('employee')

	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(employee=emp)
	elif user.role == RoleChoices.CHEF:
		try:
			chef_emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'chef has no employee record'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(employee__service_id=chef_emp.service_id)
	elif user.role not in [RoleChoices.GRH, RoleChoices.RH_SIMPLE, RoleChoices.RH_AGENT]:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	status_filter = request.query_params.get('status')
	if status_filter:
		qs = qs.filter(status=status_filter.upper())

	type_filter = request.query_params.get('type')
	if type_filter:
		qs = qs.filter(type=type_filter.upper())

	qs = qs.filter(start_date__lte=to_date, end_date__gte=from_date).order_by('start_date')

	rows = []
	for leave in qs:
		days = (leave.end_date - leave.start_date).days + 1
		rows.append([
			f"{leave.employee.first_name} {leave.employee.last_name}",
			leave.type,
			leave.start_date.isoformat(),
			leave.end_date.isoformat(),
			str(days),
			leave.status,
			leave.chef_comment or '-',
		])

	headers = ['Employee', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Chef Comment']
	title = f"Leave History Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"leave_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		return _make_export_response(
			_build_excel_report(title, headers, rows),
			filename,
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
	return _make_export_response(_build_pdf_report(title, headers, rows), filename, 'application/pdf')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_tasks_report(request):
	"""Export tasks as PDF or Excel."""
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	user = request.user
	qs = Task.objects.all().select_related('assigned_by', 'assigned_to')

	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(assigned_to=emp)
	elif user.role == RoleChoices.CHEF:
		qs = qs.filter(assigned_by=user)
	elif user.role not in [RoleChoices.GRH, RoleChoices.RH_SIMPLE, RoleChoices.RH_AGENT]:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	qs = qs.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)

	employee_id = request.query_params.get('employee_id')
	if employee_id and user.role in (RoleChoices.CHEF, RoleChoices.GRH, RoleChoices.RH_SIMPLE, RoleChoices.RH_AGENT):
		qs = qs.filter(assigned_to_id=employee_id)

	status_filter = (request.query_params.get('status') or '').lower()
	today = timezone.now().date()
	if status_filter == 'done':
		qs = qs.filter(status=Task.Status.DONE)
	elif status_filter == 'overdue':
		qs = qs.filter(status=Task.Status.TODO, due_date__lt=today)
	elif status_filter == 'assigned':
		qs = qs.filter(status=Task.Status.TODO).filter(
			Q(due_date__gte=today) | Q(due_date__isnull=True)
		)

	rows = []
	for task in qs.order_by('-created_at'):
		assigned_by = (
			f"{task.assigned_by.first_name} {task.assigned_by.last_name}".strip()
			if task.assigned_by else '-'
		)
		assigned_to = (
			f"{task.assigned_to.first_name} {task.assigned_to.last_name}"
			if task.assigned_to else '-'
		)
		if task.status == Task.Status.DONE:
			status_label = 'Done'
		elif task.due_date and task.due_date < today:
			status_label = 'Overdue'
		else:
			status_label = 'Assigned'

		rows.append([
			task.title,
			task.description or '-',
			assigned_by,
			assigned_to,
			_format_date(task.due_date),
			status_label,
			_format_date(task.completed_at.date() if task.completed_at else None),
		])

	headers = ['Title', 'Description', 'Assigned By', 'Assigned To', 'Due Date', 'Status', 'Completion Date']
	title = f"Task Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"task_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		return _make_export_response(
			_build_excel_report(title, headers, rows),
			filename,
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
	return _make_export_response(_build_pdf_report(title, headers, rows), filename, 'application/pdf')


# ──────────────────────────────────────────────────────────────────────────────
# Dashboard & KPI Export Endpoints
# ──────────────────────────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_dashboard_report(request):
	"""
	Generate a comprehensive HR report with all KPIs and charts.
	
	Request Body:
		{
			'title': 'Q1 2026 HR Report',
			'start_date': '2026-01-01',
			'end_date': '2026-03-09',
			'include_widgets': ['employee_count', 'turnover', 'absenteeism', 'evaluations'],
			'service_id': null,
			'contract_type': null,
			'include_details': true
		}
	"""
	try:
		from madar_app.services import StatisticsService, ReportFilter
		
		title = request.data.get('title', 'HR Report')
		start_date = request.data.get('start_date')
		end_date = request.data.get('end_date')
		include_widgets = request.data.get('include_widgets', ['employee_count', 'turnover', 'absenteeism', 'evaluations'])
		service_id = request.data.get('service_id')
		contract_type = request.data.get('contract_type')
		include_details = request.data.get('include_details', True)
		
		# Set defaults
		end_date_obj = datetime.fromisoformat(end_date).date() if end_date else timezone.now().date()
		start_date_obj = datetime.fromisoformat(start_date).date() if start_date else (end_date_obj - timedelta(days=90))
		
		# Create report filter
		report_filter = ReportFilter(
			start_date=start_date_obj,
			end_date=end_date_obj,
			service_id=service_id,
			contract_type=contract_type,
			employee_status='ACTIVE'
		)
		
		# Generate KPIs
		stats_service = StatisticsService()
		kpis_data = []
		
		for kpi_type in include_widgets:
			try:
				kpi_result = stats_service.calculerKPI(kpi_type, report_filter)
				
				if kpi_result:
					kpi_dict = {
						'type': kpi_result.type.value,
						'value': float(kpi_result.value),
						'unit': _get_kpi_unit_helper(kpi_type),
						'trend': kpi_result.trend.value if kpi_result.trend else None,
						'calculation_date': kpi_result.calculation_date.isoformat()
					}
					
					if include_details and kpi_result.details:
						kpi_dict['details'] = kpi_result.details
					
					kpis_data.append(kpi_dict)
			except Exception as e:
				kpis_data.append({
					'type': kpi_type,
					'error': str(e)
				})
		
		response_data = {
			'id': f"report-{int(timezone.now().timestamp())}",
			'title': title,
			'generated_at': timezone.now().isoformat(),
			'period': {
				'start_date': start_date_obj.isoformat(),
				'end_date': end_date_obj.isoformat()
			},
			'filter': {
				'service_id': service_id,
				'contract_type': contract_type,
				'employee_status': 'ACTIVE'
			},
			'kpis': kpis_data,
			'summary': {
				'total_kpis': len(kpis_data),
				'period_days': (end_date_obj - start_date_obj).days,
				'generated_by': request.user.email,
				'generated_at': timezone.now().isoformat()
			}
		}
		
		return Response(response_data)
	
	except ValueError as e:
		return Response({'error': f'Invalid parameter: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
	except Exception as e:
		return Response({'error': f'Failed to generate report: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_dashboard_report(request):
	"""
	Export the dashboard/KPI report as PDF, Excel, or CSV.
	
	Query Parameters:
		- format: 'pdf' | 'excel' | 'csv'
	"""
	try:
		from madar_app.services import ExportService
		
		export_format = request.query_params.get('format', 'pdf').lower()
		report_data = request.data.get('report')
		
		if not report_data:
			return Response({'error': 'Report data is required'}, status=status.HTTP_400_BAD_REQUEST)
		
		if export_format not in ['pdf', 'excel', 'csv']:
			return Response({'error': f'Invalid format: {export_format}'}, status=status.HTTP_400_BAD_REQUEST)
		
		export_service = ExportService()
		export_file = None
		
		if export_format == 'pdf':
			export_file = export_service.export_pdf([report_data])
			content_type = 'application/pdf'
		elif export_format == 'excel':
			export_file = export_service.export_excel([report_data])
			content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		else:  # csv
			export_file = export_service.export_csv([report_data])
			content_type = 'text/csv'
		
		if not export_file:
			return Response({'error': f'Failed to export to {export_format}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
		
		response = HttpResponse(export_file.content, content_type=content_type)
		response['Content-Disposition'] = f'attachment; filename="{export_file.filename}"'
		
		return response
	
	except Exception as e:
		return Response({'error': f'Failed to export: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _get_kpi_unit_helper(kpi_type):
	"""Helper: get the unit of measurement for a KPI type."""
	units = {
		'employee_count': 'people',
		'turnover': '%',
		'absenteeism': '%',
		'evaluations': 'score'
	}
	return units.get(kpi_type, 'value')
