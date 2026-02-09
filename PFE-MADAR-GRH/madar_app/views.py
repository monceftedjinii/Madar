from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from .permissions import IsGRH
from rest_framework.permissions import IsAuthenticated
from .scopes import employee_queryset_for
from .models import Task, Employee, User, RoleChoices, Department
from rest_framework import status
from rest_framework.parsers import JSONParser
from .permissions import IsChef
from .permissions import IsEmployee
from .permissions import IsRHSimple, IsRHSenior, CanUploadDocument, CanValidateDocument
from .models import Attendance, LeaveRequest, AbsenceWarning, DisciplineFlag, Notification
from .models import Document, DocumentHistory, DocumentType
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q
from datetime import date, datetime, timedelta
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def notify(user, title, message):
	"""Helper to create a notification."""
	return Notification.objects.create(user=user, title=title, message=message)


@api_view(['GET'])
def ping(request):
	return Response({'ping': 'pong'})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsGRH])
def rbac_test(request):
	# returns the role of the authenticated user; protected to GRH only
	role = getattr(request.user, 'role', None)
	return Response({'ok': True, 'role': role})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def whoami(request):
	user = request.user
	return Response({
		'id': user.id,
		'email': user.email,
		'role': getattr(user, 'role', None),
	})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employees_list(request):
	qs = employee_queryset_for(request.user)
	data = [
		{
			'id': e.id,
			'first_name': e.first_name,
			'last_name': e.last_name,
			'email': e.email,
			'department': {
				'id': e.department.id,
				'name': e.department.name,
			} if e.department else None,
		}
		for e in qs.order_by('id')
	]
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def departments_list(request):
	data = [
		{'id': d.id, 'name': d.name}
		for d in Department.objects.order_by('name')
	]
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsChef])
def create_task(request):
	# Chef assigns a task to an employee in his department only
	import logging
	logger = logging.getLogger(__name__)
	
	data = request.data
	logger.info(f'create_task request from {request.user.email} with data: {data}')
	
	title = data.get('title')
	if not title:
		logger.warning(f'create_task: title is required')
		return Response({'detail': 'title is required'}, status=status.HTTP_400_BAD_REQUEST)
	
	assigned_to_id = data.get('assigned_to')
	if not assigned_to_id:
		logger.warning(f'create_task: assigned_to is required')
		return Response({'detail': 'assigned_to is required'}, status=status.HTTP_400_BAD_REQUEST)
	
	try:
		emp = Employee.objects.get(id=assigned_to_id)
		logger.info(f'create_task: found employee {emp.email}')
	except Employee.DoesNotExist:
		logger.warning(f'create_task: employee {assigned_to_id} not found')
		return Response({'detail': 'assigned_to not found'}, status=status.HTTP_400_BAD_REQUEST)

	# ensure chef's department matches emp.department
	try:
		chef_emp = Employee.objects.get(email=request.user.email)
		logger.info(f'create_task: chef {chef_emp.email} found in dept {chef_emp.department.name}')
	except Employee.DoesNotExist:
		logger.warning(f'create_task: chef {request.user.email} has no employee record')
		return Response({'detail': 'chef has no employee record'}, status=status.HTTP_400_BAD_REQUEST)

	if emp.department_id != chef_emp.department_id:
		logger.warning(f'create_task: chef dept {chef_emp.department_id} != employee dept {emp.department_id}')
		return Response({'detail': 'cannot assign outside your department'}, status=status.HTTP_403_FORBIDDEN)

	task = Task.objects.create(
		title=title,
		description=data.get('description', ''),
		due_date=data.get('due_date', None),
		assigned_to=emp,
		assigned_by=request.user,
	)
	logger.info(f'create_task: task {task.id} created and assigned to {emp.email}')
	return Response({'id': task.id}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_tasks(request):
	# employee view of their tasks
	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response([], status=status.HTTP_200_OK)
	qs = Task.objects.filter(assigned_to=emp).order_by('-created_at')
	data = [{
		'id': t.id,
		'title': t.title,
		'description': t.description,
		'status': t.status,
		'due_date': t.due_date,
		'created_at': t.created_at,
		'completed_at': t.completed_at,
		'assigned_by': {
			'id': t.assigned_by.id,
			'email': t.assigned_by.email,
			'first_name': t.assigned_by.first_name,
			'last_name': t.assigned_by.last_name,
		} if t.assigned_by else None
	} for t in qs]
	return Response(data)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def mark_task_done(request, pk):
	from django.utils import timezone
	import logging
	logger = logging.getLogger(__name__)
	
	try:
		task = Task.objects.get(id=pk)
	except Task.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	# only assigned employee can mark as done
	if task.assigned_to.email != request.user.email:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	task.status = Task.Status.DONE
	task.completed_at = timezone.now()
	task.save()
	
	# Notify the chef who assigned this task
	if task.assigned_by:
		emp = task.assigned_to
		chef_name = f"{task.assigned_by.first_name} {task.assigned_by.last_name}".strip() or task.assigned_by.email
		notify(
			task.assigned_by,
			title=f"Task Completed",
			message=f"{emp.first_name} {emp.last_name} marked '{task.title}' as done"
		)
		logger.info(f'mark_task_done: notified chef {task.assigned_by.email} that task {task.id} was completed')
	
	return Response({'ok': True})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsChef])
def chef_tasks(request):
	# Chef view of all tasks assigned by this chef
	import logging
	logger = logging.getLogger(__name__)
	
	qs = Task.objects.filter(assigned_by=request.user).order_by('-created_at')
	logger.info(f'chef_tasks: fetching tasks for chef {request.user.email}, found {qs.count()} tasks')
	
	data = [{
		'id': t.id,
		'title': t.title,
		'description': t.description,
		'status': t.status,
		'due_date': t.due_date,
		'created_at': t.created_at,
		'completed_at': t.completed_at,
		'employee': {
			'id': t.assigned_to.id,
			'email': t.assigned_to.email,
			'first_name': t.assigned_to.first_name,
			'last_name': t.assigned_to.last_name,
			'department': {
				'id': t.assigned_to.department.id,
				'name': t.assigned_to.department.name
			} if t.assigned_to.department else None
		},
		'assigned_by': {
			'id': t.assigned_by.id,
			'email': t.assigned_by.email,
			'first_name': t.assigned_by.first_name,
			'last_name': t.assigned_by.last_name,
		} if t.assigned_by else None
	} for t in qs]
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsEmployee])
def attendance_check_in(request):
	pin = request.data.get('pin')
	if not pin:
		return Response({'detail': 'pin is required'}, status=status.HTTP_400_BAD_REQUEST)
	if not str(pin).isdigit() or len(str(pin)) != 4:
		return Response({'detail': 'pin must be exactly 4 digits'}, status=status.HTTP_400_BAD_REQUEST)
	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
	if not emp.attendance_pin:
		return Response({'detail': 'pin not set'}, status=status.HTTP_400_BAD_REQUEST)
	if str(pin) != emp.attendance_pin:
		return Response({'detail': 'Invalid PIN'}, status=status.HTTP_403_FORBIDDEN)

	now = timezone.now()
	today = now.date()

	att, created = Attendance.objects.get_or_create(employee=emp, date=today)
	if att.check_in_time:
		return Response({'detail': 'already checked in'}, status=status.HTTP_400_BAD_REQUEST)

	att.check_in_time = now.time()
	att.save()
	return Response({'id': att.id, 'check_in_time': att.check_in_time.isoformat()}, status=(status.HTTP_201_CREATED if created else status.HTTP_200_OK))


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsEmployee])
def attendance_check_out(request):
	pin = request.data.get('pin')
	if not pin:
		return Response({'detail': 'pin is required'}, status=status.HTTP_400_BAD_REQUEST)
	if not str(pin).isdigit() or len(str(pin)) != 4:
		return Response({'detail': 'pin must be exactly 4 digits'}, status=status.HTTP_400_BAD_REQUEST)
	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
	if not emp.attendance_pin:
		return Response({'detail': 'pin not set'}, status=status.HTTP_400_BAD_REQUEST)
	if str(pin) != emp.attendance_pin:
		return Response({'detail': 'Invalid PIN'}, status=status.HTTP_403_FORBIDDEN)

	now = timezone.now()
	today = now.date()

	try:
		att = Attendance.objects.get(employee=emp, date=today)
	except Attendance.DoesNotExist:
		return Response({'detail': 'no check-in found for today'}, status=status.HTTP_400_BAD_REQUEST)

	if not att.check_in_time:
		return Response({'detail': 'no check-in found for today'}, status=status.HTTP_400_BAD_REQUEST)
	if att.check_out_time:
		return Response({'detail': 'already checked out'}, status=status.HTTP_400_BAD_REQUEST)

	att.check_out_time = now.time()
	att.save()
	return Response({'id': att.id, 'check_out_time': att.check_out_time.isoformat()})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsEmployee])
def attendance_me(request):
	# parse from/to
	qfrom = request.query_params.get('from')
	qto = request.query_params.get('to')
	today = timezone.now().date()
	if qfrom:
		from_date = datetime.fromisoformat(qfrom).date()
	else:
		from_date = today.replace(day=1)
	if qto:
		to_date = datetime.fromisoformat(qto).date()
	else:
		to_date = today

	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response([], status=status.HTTP_200_OK)

	qs = Attendance.objects.filter(employee=emp, date__gte=from_date, date__lte=to_date).order_by('date')
	data = [
		{'date': a.date.isoformat(), 'check_in_time': a.check_in_time.isoformat() if a.check_in_time else None, 'check_out_time': a.check_out_time.isoformat() if a.check_out_time else None}
		for a in qs
	]
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsEmployee])
def create_leave(request):
	# employee creates own leave request
	data = request.data
	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)

	ltype = data.get('type')
	start_date = data.get('start_date')
	end_date = data.get('end_date')
	reason = data.get('reason', '')

	# basic required fields
	if not start_date or not end_date:
		return Response({'detail': 'start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

	# parse dates
	try:
		sd = date.fromisoformat(start_date)
		ed = date.fromisoformat(end_date)
	except Exception:
		return Response({'detail': 'invalid date format, use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

	# end must be >= start
	if ed < sd:
		return Response({'detail': 'end_date must be the same or after start_date'}, status=status.HTTP_400_BAD_REQUEST)

	# sick leave requires attachment
	if ltype == LeaveRequest.LeaveType.SICK and not request.FILES.get('attachment'):
		return Response({'detail': 'attachment required for sick leave'}, status=status.HTTP_400_BAD_REQUEST)

	# Business rule: block if employee has any PENDING request or ongoing ACCEPTED leave
	today = date.today()
	blocked = LeaveRequest.objects.filter(
		employee=emp,
		# PENDING: any
		status=LeaveRequest.Status.PENDING
	).exists() or LeaveRequest.objects.filter(
		employee=emp,
		status=LeaveRequest.Status.ACCEPTED,
		end_date__gte=today
	).exists()
	if blocked:
		return Response({'detail': "You can’t submit a new leave request while you have a pending request or an ongoing approved leave."}, status=status.HTTP_400_BAD_REQUEST)

	leave = LeaveRequest.objects.create(
		employee=emp,
		start_date=sd,
		end_date=ed,
		type=ltype or LeaveRequest.LeaveType.ANNUAL,
		reason=reason,
		attachment=request.FILES.get('attachment') if request.FILES.get('attachment') else None,
		status=LeaveRequest.Status.PENDING,
	)

	# Notify chef(s) in the same department
	chef_emails = Employee.objects.filter(department=emp.department).values_list('email', flat=True)
	chef_users = User.objects.filter(role=RoleChoices.CHEF, email__in=chef_emails)
	for chef_user in chef_users:
		notify(chef_user, 'New leave request', f'{emp.first_name} {emp.last_name} requested leave from {sd} to {ed}.')

	return Response({'id': leave.id}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsEmployee])
def my_leaves(request):
	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response([], status=status.HTTP_200_OK)
	qs = LeaveRequest.objects.filter(employee=emp).order_by('-created_at')
	data = [
		{
			'id': l.id,
			'start_date': l.start_date.isoformat(),
			'end_date': l.end_date.isoformat(),
			'type': l.type,
			'status': l.status,
			'reason': l.reason,
			'chef_comment': l.chef_comment,
			'decided_by': l.decided_by.email if l.decided_by else None,
			'decided_at': l.decided_at.isoformat() if l.decided_at else None,
		}
		for l in qs
	]
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsRHSimple])
def absences_yesterday(request):
	# list employees absent yesterday (no attendance, no accepted leave)
	yesterday = (timezone.now().date() - timezone.timedelta(days=1))
	attended_ids = Attendance.objects.filter(date=yesterday).values_list('employee_id', flat=True)
	onleave_ids = LeaveRequest.objects.filter(status=LeaveRequest.Status.ACCEPTED, start_date__lte=yesterday, end_date__gte=yesterday).values_list('employee_id', flat=True)

	qs = Employee.objects.exclude(id__in=attended_ids).exclude(id__in=onleave_ids)
	data = [
		{'id': e.id, 'full_name': f"{e.first_name} {e.last_name}", 'department': e.department.name if e.department else None}
		for e in qs.order_by('id')
	]
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsRHSimple])
def create_warning(request):
	emp_id = request.data.get('employee_id')
	date_str = request.data.get('date')
	comment = request.data.get('comment', '')
	if not emp_id or not date_str:
		return Response({'detail': 'employee_id and date are required'}, status=status.HTTP_400_BAD_REQUEST)
	try:
		dt = date.fromisoformat(date_str)
	except Exception:
		return Response({'detail': 'invalid date format, use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		emp = Employee.objects.get(id=emp_id)
	except Employee.DoesNotExist:
		return Response({'detail': 'employee not found'}, status=status.HTTP_400_BAD_REQUEST)

	if AbsenceWarning.objects.filter(employee=emp, date=dt).exists():
		return Response({'detail': 'warning for this employee and date already exists'}, status=status.HTTP_400_BAD_REQUEST)

	aw = AbsenceWarning.objects.create(employee=emp, date=dt, comment=comment, issued_by=request.user)

	month_start = dt.replace(day=1)
	flag, created = DisciplineFlag.objects.get_or_create(employee=emp, month=month_start, defaults={'warning_count': 0})
	flag.warning_count = flag.warning_count + 1
	flag.save()

	# notify RH_SENIOR if flag reaches 3
	if flag.warning_count >= 3:
		rh_senior_users = User.objects.filter(role=RoleChoices.RH_SENIOR)
		for rh_user in rh_senior_users:
			notify(rh_user, 'Discipline Flag', f'Employee {emp.first_name} {emp.last_name} has reached {flag.warning_count} warnings in the current month.')

	return Response({'id': aw.id, 'warning_count': flag.warning_count}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsRHSenior])
def discipline_flags(request):
	today = timezone.now().date()
	month_start = today.replace(day=1)
	qs = DisciplineFlag.objects.filter(month=month_start, warning_count__gte=3).order_by('-warning_count')
	data = [
		{
			'employee_id': f.employee.id,
			'employee_email': f.employee.email,
			'warning_count': f.warning_count,
			'month': f.month.isoformat(),
		}
		for f in qs
	]
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_notifications(request):
	# list user's own notifications (latest first)
	qs = Notification.objects.filter(user=request.user).order_by('-created_at')
	data = [
		{
			'id': n.id,
			'title': n.title,
			'message': n.message,
			'is_read': n.is_read,
			'created_at': n.created_at.isoformat(),
		}
		for n in qs
	]
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_notification_read(request, pk):
	try:
		notif = Notification.objects.get(id=pk)
	except Notification.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)
	
	if notif.user != request.user:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)
	
	notif.is_read = True
	notif.save()
	return Response({'id': notif.id, 'is_read': True})


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsChef])
def department_pending_leaves(request):
	# Chef lists all leaves in his department (history + pending)
	try:
		chef_emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response(
			{'detail': 'Chef has no Employee record / department assigned'},
			status=status.HTTP_400_BAD_REQUEST
		)
	qs = LeaveRequest.objects.filter(employee__department=chef_emp.department).order_by('-created_at')
	print(f"department_pending_leaves dept_id={chef_emp.department_id} total_count={qs.count()}")
	data = [
		{
			'id': l.id,
			'employee_email': l.employee.email,
			'employee': {
				'email': l.employee.email,
				'first_name': l.employee.first_name,
				'last_name': l.employee.last_name,
			},
			'start_date': l.start_date.isoformat(),
			'end_date': l.end_date.isoformat(),
			'type': l.type,
			'reason': l.reason,
			'attachment': request.build_absolute_uri(l.attachment.url) if l.attachment else None,
			'status': l.status,
		}
		for l in qs
	]
	return Response(data)


def _chef_decide_common(request, pk, accept=True):
	try:
		lr = LeaveRequest.objects.get(id=pk)
	except LeaveRequest.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	if lr.status != LeaveRequest.Status.PENDING:
		return Response({'detail': 'leave not pending'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		chef_emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response({'detail': 'chef has no employee record'}, status=status.HTTP_400_BAD_REQUEST)

	if lr.employee.department_id != chef_emp.department_id:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	comment = request.data.get('comment', '')
	lr.chef_comment = comment
	lr.decided_by = request.user
	lr.decided_at = timezone.now()
	lr.status = LeaveRequest.Status.ACCEPTED if accept else LeaveRequest.Status.REFUSED
	lr.save()

	# notify employee
	emp_user = None
	try:
		emp_user = User.objects.get(email=lr.employee.email)
	except:
		pass
	if emp_user:
		status_label = 'approved' if accept else 'rejected'
		notify(emp_user, f'Leave {status_label}', f'Your leave request from {lr.start_date} to {lr.end_date} has been {status_label}.')

	return Response({'id': lr.id, 'status': lr.status})


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsChef])
def approve_leave(request, pk):
	return _chef_decide_common(request, pk, accept=True)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsChef])
def reject_leave(request, pk):
	return _chef_decide_common(request, pk, accept=False)

# Document Helper
def create_doc_history(document, action, by_user, note='', parent=None):
	"""Helper to create a document history entry."""
	return DocumentHistory.objects.create(
		document=document,
		parent=parent,
		action=action,
		by_user=by_user,
		note=note
	)


# Document Endpoints
@api_view(['POST'])
@permission_classes([IsAuthenticated, CanUploadDocument])
def upload_document(request):
	"""Upload a new document (DRAFT status)."""
	title = request.data.get('title')
	file_obj = request.FILES.get('file')
	if not all([title, file_obj]):
		return Response({'detail': 'title and file required'}, status=status.HTTP_400_BAD_REQUEST)

	doc_type_id = request.data.get('doc_type')
	doc_type_name = request.data.get('type')
	category = (request.data.get('category') or DocumentType.Category.INTERNAL).upper()

	if doc_type_id:
		try:
			doc_type = DocumentType.objects.get(id=doc_type_id)
		except DocumentType.DoesNotExist:
			return Response({'detail': 'doc_type not found'}, status=status.HTTP_400_BAD_REQUEST)
	else:
		if not doc_type_name:
			return Response({'detail': 'type is required'}, status=status.HTTP_400_BAD_REQUEST)
		doc_type, _ = DocumentType.objects.get_or_create(
			name=doc_type_name,
			defaults={'category': category},
		)
		if doc_type.category != category:
			doc_type.category = category
			doc_type.save()

	def _resolve_department(value):
		if not value:
			return None
		if str(value).isdigit():
			return Department.objects.filter(id=int(value)).first()
		return Department.objects.filter(name=value).first()

	source_dept = _resolve_department(request.data.get('source_department'))
	if not source_dept:
		try:
			user_emp = Employee.objects.get(email=request.user.email)
			source_dept = user_emp.department
		except Employee.DoesNotExist:
			return Response({'detail': 'source_department is required'}, status=status.HTTP_400_BAD_REQUEST)

	target_dept = _resolve_department(request.data.get('target_department'))
	if request.user.role == RoleChoices.CHEF and not target_dept:
		return Response({'detail': 'target_department is required'}, status=status.HTTP_400_BAD_REQUEST)

	# Permission check: employee/chef must upload for their own department
	if request.user.role in [RoleChoices.EMPLOYEE, RoleChoices.CHEF]:
		try:
			emp = Employee.objects.get(email=request.user.email)
			if source_dept and emp.department_id != source_dept.id:
				return Response({'detail': 'can only upload for own department'}, status=status.HTTP_403_FORBIDDEN)
		except Employee.DoesNotExist:
			return Response({'detail': 'user has no employee record'}, status=status.HTTP_403_FORBIDDEN)

	# RH_SIMPLE can only upload RH documents
	if request.user.role == RoleChoices.RH_SIMPLE and doc_type.category != DocumentType.Category.RH:
		return Response({'detail': 'RH_SIMPLE can only upload RH documents'}, status=status.HTTP_403_FORBIDDEN)

	doc = Document.objects.create(
		title=title,
		doc_type=doc_type,
		file=file_obj,
		source_department=source_dept,
		target_department=target_dept,
		created_by=request.user,
		status=Document.Status.DRAFT
	)

	create_doc_history(doc, DocumentHistory.Action.CREATED, request.user)

	file_url = request.build_absolute_uri(doc.file.url) if doc.file else None
	created_by_name = None
	if doc.created_by:
		name = f"{doc.created_by.first_name} {doc.created_by.last_name}".strip()
		created_by_name = name or doc.created_by.email
	return Response({
		'id': doc.id,
		'title': doc.title,
		'doc_type': doc.doc_type.name,
		'doc_type_category': doc.doc_type.category,
		'status': doc.status,
		'source_department': doc.source_department.name if doc.source_department else None,
		'target_department': doc.target_department.name if doc.target_department else None,
		'created_by': doc.created_by.email if doc.created_by else None,
		'created_by_name': created_by_name,
		'created_at': doc.created_at.isoformat() if doc.created_at else None,
		'sent_at': doc.sent_at.isoformat() if doc.sent_at else None,
		'file_url': file_url,
	}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_document(request, pk):
	"""Send a document (change status from DRAFT -> SENT)."""
	try:
		doc = Document.objects.get(id=pk)
	except Document.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	if doc.status != Document.Status.DRAFT:
		return Response({'detail': 'can only send DRAFT documents'}, status=status.HTTP_400_BAD_REQUEST)

	if not doc.target_department_id:
		return Response({'detail': 'target_department is required to send'}, status=status.HTTP_400_BAD_REQUEST)

	# Permission: only creator or chef of source department
	is_creator = doc.created_by_id == request.user.id
	is_chef_of_dept = False
	if request.user.role == RoleChoices.CHEF:
		try:
			chef_emp = Employee.objects.get(email=request.user.email)
			is_chef_of_dept = chef_emp.department_id == doc.source_department_id
		except Employee.DoesNotExist:
			pass

	if not (is_creator or is_chef_of_dept):
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	doc.status = Document.Status.SENT
	doc.sent_at = timezone.now()
	doc.save()
	create_doc_history(doc, DocumentHistory.Action.SENT, request.user)

	if doc.target_department_id:
		dept_emps = Employee.objects.filter(department_id=doc.target_department_id)
		for emp in dept_emps:
			try:
				user = User.objects.get(email=emp.email)
			except User.DoesNotExist:
				continue
			notify(
				user,
				title='New document received',
				message=f"{doc.title} was sent to your department."
			)

	return Response({'id': doc.id, 'status': doc.status, 'sent_at': doc.sent_at.isoformat()})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_documents_scoped(request):
	"""List documents scoped to user's role and department."""
	if request.user.role == RoleChoices.GRH:
		# GRH sees all
		qs = Document.objects.all()
	elif request.user.role == RoleChoices.RH_SENIOR:
		# RH_SENIOR sees all (for validation)
		qs = Document.objects.all()
	elif request.user.role == RoleChoices.RH_SIMPLE:
		# RH_SIMPLE sees only RH documents they created
		qs = Document.objects.filter(created_by=request.user, doc_type__category=DocumentType.Category.RH)
	elif request.user.role == RoleChoices.CHEF:
		# Chef sees documents of his department (as source or target)
		try:
			chef_emp = Employee.objects.get(email=request.user.email)
			from django.db.models import Q
			qs = Document.objects.filter(
				Q(source_department_id=chef_emp.department_id) |
				Q(target_department_id=chef_emp.department_id)
			)
		except Employee.DoesNotExist:
			qs = Document.objects.none()
	elif request.user.role == RoleChoices.EMPLOYEE:
		# Employee sees documents they created + documents sent to their department
		try:
			emp = Employee.objects.get(email=request.user.email)
			from django.db.models import Q
			qs = Document.objects.filter(
				Q(created_by=request.user) |
				Q(target_department_id=emp.department_id, status__in=[Document.Status.SENT, Document.Status.VALIDATED, Document.Status.ARCHIVED]) |
				Q(source_department_id=emp.department_id, status__in=[Document.Status.SENT, Document.Status.VALIDATED, Document.Status.ARCHIVED])
			)
		except Employee.DoesNotExist:
			qs = Document.objects.filter(created_by=request.user)
	else:
		qs = Document.objects.none()

	data = []
	for d in qs.order_by('-created_at'):
		created_by_name = None
		if d.created_by:
			name = f"{d.created_by.first_name} {d.created_by.last_name}".strip()
			created_by_name = name or d.created_by.email
		data.append({
			'id': d.id,
			'title': d.title,
			'doc_type': d.doc_type.name,
			'doc_type_category': d.doc_type.category,
			'status': d.status,
			'source_department': d.source_department.name if d.source_department else None,
			'target_department': d.target_department.name if d.target_department else None,
			'created_by': d.created_by.email if d.created_by else None,
			'created_by_name': created_by_name,
			'created_at': d.created_at.isoformat() if d.created_at else None,
			'sent_at': d.sent_at.isoformat() if d.sent_at else None,
			'file_url': request.build_absolute_uri(d.file.url) if d.file else None,
		})
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def documents_feed(request):
	"""Employee feed: documents sent to the employee's department."""
	if request.user.role != RoleChoices.EMPLOYEE:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	try:
		emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response([], status=status.HTTP_200_OK)

	qs = Document.objects.filter(
		target_department_id=emp.department_id,
		status=Document.Status.SENT
	)

	data = []
	for d in qs.order_by('-sent_at', '-created_at'):
		created_by_name = None
		if d.created_by:
			name = f"{d.created_by.first_name} {d.created_by.last_name}".strip()
			created_by_name = name or d.created_by.email
		data.append({
			'id': d.id,
			'title': d.title,
			'doc_type': d.doc_type.name,
			'doc_type_category': d.doc_type.category,
			'status': d.status,
			'source_department': d.source_department.name if d.source_department else None,
			'target_department': d.target_department.name if d.target_department else None,
			'created_by': d.created_by.email if d.created_by else None,
			'created_by_name': created_by_name,
			'created_at': d.created_at.isoformat() if d.created_at else None,
			'sent_at': d.sent_at.isoformat() if d.sent_at else None,
			'file_url': request.build_absolute_uri(d.file.url) if d.file else None,
		})
	return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def documents_mine(request):
	"""Chef history: documents posted from the chef's department."""
	if request.user.role != RoleChoices.CHEF:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	try:
		chef_emp = Employee.objects.get(email=request.user.email)
	except Employee.DoesNotExist:
		return Response([], status=status.HTTP_200_OK)

	qs = Document.objects.filter(source_department_id=chef_emp.department_id)

	data = []
	for d in qs.order_by('-created_at'):
		created_by_name = None
		if d.created_by:
			name = f"{d.created_by.first_name} {d.created_by.last_name}".strip()
			created_by_name = name or d.created_by.email
		data.append({
			'id': d.id,
			'title': d.title,
			'doc_type': d.doc_type.name,
			'doc_type_category': d.doc_type.category,
			'status': d.status,
			'source_department': d.source_department.name if d.source_department else None,
			'target_department': d.target_department.name if d.target_department else None,
			'created_by': d.created_by.email if d.created_by else None,
			'created_by_name': created_by_name,
			'created_at': d.created_at.isoformat() if d.created_at else None,
			'sent_at': d.sent_at.isoformat() if d.sent_at else None,
			'file_url': request.build_absolute_uri(d.file.url) if d.file else None,
		})
	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def comment_document(request, pk):
	"""Add a comment to a document (stored as COMMENTED history)."""
	try:
		doc = Document.objects.get(id=pk)
	except Document.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	def _can_access_comments(user, document):
		# creator, chef of source/target, RH_SENIOR, GRH, or employee in source/target dept
		if user.id == document.created_by_id:
			return True
		if user.role == RoleChoices.CHEF:
			try:
				chef_emp = Employee.objects.get(email=user.email)
				return (
					chef_emp.department_id == document.source_department_id or
					(document.target_department_id and chef_emp.department_id == document.target_department_id)
				)
			except Employee.DoesNotExist:
				return False
		if user.role in [RoleChoices.RH_SENIOR, RoleChoices.GRH]:
			return True
		if user.role == RoleChoices.EMPLOYEE:
			try:
				emp = Employee.objects.get(email=user.email)
				return (
					emp.department_id == document.source_department_id or
					(document.target_department_id and emp.department_id == document.target_department_id)
				)
			except Employee.DoesNotExist:
				return False
		return False

	if not _can_access_comments(request.user, doc):
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	comment = request.data.get('comment', '')
	if not comment:
		return Response({'detail': 'comment required'}, status=status.HTTP_400_BAD_REQUEST)

	parent = None
	parent_id = request.data.get('parent_id')
	if parent_id:
		try:
			parent = DocumentHistory.objects.get(
				id=parent_id,
				document=doc,
				action=DocumentHistory.Action.COMMENTED
			)
		except DocumentHistory.DoesNotExist:
			return Response({'detail': 'parent comment not found'}, status=status.HTTP_400_BAD_REQUEST)

	create_doc_history(doc, DocumentHistory.Action.COMMENTED, request.user, note=comment, parent=parent)
	return Response({'detail': 'comment added'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def document_comments(request, pk):
	"""List comments for a document."""
	try:
		doc = Document.objects.get(id=pk)
	except Document.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	def _can_access_comments(user, document):
		if user.id == document.created_by_id:
			return True
		if user.role == RoleChoices.CHEF:
			try:
				chef_emp = Employee.objects.get(email=user.email)
				return (
					chef_emp.department_id == document.source_department_id or
					(document.target_department_id and chef_emp.department_id == document.target_department_id)
				)
			except Employee.DoesNotExist:
				return False
		if user.role in [RoleChoices.RH_SENIOR, RoleChoices.GRH]:
			return True
		if user.role == RoleChoices.EMPLOYEE:
			try:
				emp = Employee.objects.get(email=user.email)
				return (
					emp.department_id == document.source_department_id or
					(document.target_department_id and emp.department_id == document.target_department_id)
				)
			except Employee.DoesNotExist:
				return False
		return False

	if not _can_access_comments(request.user, doc):
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	comments = DocumentHistory.objects.filter(
		document=doc,
		action=DocumentHistory.Action.COMMENTED
	).select_related('by_user', 'parent').order_by('created_at')

	comment_map = {}
	ordered = []
	for c in comments:
		by_user_name = None
		if c.by_user:
			name = f"{c.by_user.first_name} {c.by_user.last_name}".strip()
			by_user_name = name or c.by_user.email
		item = {
			'id': c.id,
			'note': c.note,
			'by_user': c.by_user.email if c.by_user else None,
			'by_user_name': by_user_name,
			'created_at': c.created_at.isoformat() if c.created_at else None,
			'parent_id': c.parent_id,
			'replies': [],
		}
		comment_map[c.id] = item
		ordered.append(item)

	data = []
	for item in ordered:
		if item['parent_id'] and item['parent_id'] in comment_map:
			comment_map[item['parent_id']]['replies'].append(item)
		else:
			data.append(item)

	return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, CanValidateDocument])
def validate_document(request, pk):
	"""Validate a document (RH_SENIOR/GRH only, status -> VALIDATED)."""
	try:
		doc = Document.objects.get(id=pk)
	except Document.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	if doc.status == Document.Status.VALIDATED:
		return Response({'detail': 'already validated'}, status=status.HTTP_400_BAD_REQUEST)

	if doc.status in [Document.Status.ARCHIVED]:
		return Response({'detail': 'cannot validate archived documents'}, status=status.HTTP_400_BAD_REQUEST)

	doc.status = Document.Status.VALIDATED
	doc.validated_by = request.user
	doc.validated_at = timezone.now()
	doc.save()

	create_doc_history(doc, DocumentHistory.Action.VALIDATED, request.user)
	return Response({'id': doc.id, 'status': doc.status})


@api_view(['POST'])
@permission_classes([IsAuthenticated, CanValidateDocument])
def archive_document(request, pk):
	"""Archive a document (RH_SENIOR/GRH only, status -> ARCHIVED)."""
	try:
		doc = Document.objects.get(id=pk)
	except Document.DoesNotExist:
		return Response({'detail': 'not found'}, status=status.HTTP_404_NOT_FOUND)

	if doc.status == Document.Status.ARCHIVED:
		return Response({'detail': 'already archived'}, status=status.HTTP_400_BAD_REQUEST)

	doc.status = Document.Status.ARCHIVED
	doc.save()

	create_doc_history(doc, DocumentHistory.Action.ARCHIVED, request.user)
	return Response({'id': doc.id, 'status': doc.status})


# Reports & Statistics Helpers
def _parse_date_str(value):
	if not value:
		return None
	try:
		return date.fromisoformat(value)
	except ValueError:
		pass
	if '/' in value:
		parts = value.split('/')
		if len(parts) == 3:
			day, month, year = parts
			return date(int(year), int(month), int(day))
	raise ValueError('invalid date format')


def parse_date_range(request):
	"""Parse ?from= and ?to= query params. Default to current month."""
	from_param = request.query_params.get('from')
	to_param = request.query_params.get('to')

	today = date.today()

	if from_param:
		from_date = _parse_date_str(from_param)
	else:
		from_date = date(today.year, today.month, 1)

	if to_param:
		to_date = _parse_date_str(to_param)
	else:
		# Default to end of current month
		if today.month == 12:
			to_date = date(today.year + 1, 1, 1) - timedelta(days=1)
		else:
			to_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

	return from_date, to_date


EXPORT_LATE_THRESHOLD = datetime.strptime('09:00', '%H:%M').time()


def _make_export_response(buffer, filename, content_type):
	response = HttpResponse(buffer.getvalue(), content_type=content_type)
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	return response


def _build_excel_report(title, headers, rows):
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = 'Report'

	worksheet.append([title])
	worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
	worksheet['A1'].font = Font(bold=True, size=14)
	worksheet['A1'].alignment = Alignment(horizontal='center')

	worksheet.append(headers)
	for cell in worksheet[2]:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal='center')

	for row in rows:
		worksheet.append(row)

	from openpyxl.utils import get_column_letter
	for col_idx in range(1, len(headers) + 1):
		max_len = 0
		for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
			for value in cell:
				if value is not None:
					max_len = max(max_len, len(str(value)))
		col_letter = get_column_letter(col_idx)
		worksheet.column_dimensions[col_letter].width = min(max_len + 2, 40)

	output = BytesIO()
	workbook.save(output)
	output.seek(0)
	return output


def _build_pdf_report(title, headers, rows):
	output = BytesIO()
	doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=24, bottomMargin=24)
	styles = getSampleStyleSheet()
	story = [Paragraph(title, styles['Title']), Spacer(1, 12)]

	data = [headers] + rows
	table = Table(data, repeatRows=1)
	table.setStyle(TableStyle([
		('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
		('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
		('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
		('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
		('ALIGN', (0, 0), (-1, 0), 'CENTER'),
		('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
	]))

	story.append(table)
	doc.build(story)
	output.seek(0)
	return output


def _resolve_employee_scope(user):
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return None
		return Employee.objects.filter(id=emp.id)
	if user.role == RoleChoices.CHEF:
		try:
			chef_emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return None
		return Employee.objects.filter(department_id=chef_emp.department_id)
	if user.role == RoleChoices.GRH:
		return Employee.objects.all()
	return None


def _format_time(value):
	if not value:
		return '-'
	return value.strftime('%H:%M')


def _format_date(value):
	if not value:
		return '-'
	return value.isoformat()


def _format_hours(value):
	return f"{value:.2f}"


def get_employee_scope(user):
	"""Return department for CHEF, None for global roles."""
	if user.role == RoleChoices.CHEF:
		try:
			emp = Employee.objects.get(email=user.email)
			return emp.department_id
		except Employee.DoesNotExist:
			return None
	elif user.role in [RoleChoices.RH_SIMPLE, RoleChoices.RH_SENIOR, RoleChoices.GRH]:
		return None  # Global
	return None  # EMPLOYEE is handled separately


def get_employees_count(user, from_date, to_date):
	"""Count employees visible to user."""
	if user.role == RoleChoices.EMPLOYEE:
		return 1
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			return Employee.objects.filter(department_id=dept_id).count()
		return 0
	else:  # GRH, RH_*
		return Employee.objects.count()


def get_attendance_count(user, from_date, to_date):
	"""Count attendance records in date range."""
	from django.db.models import Q
	
	qs = Attendance.objects.filter(date__gte=from_date, date__lte=to_date)
	
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			qs = qs.none()
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			qs = qs.filter(employee__department_id=dept_id)
		else:
			qs = qs.none()
	# else: GRH, RH_* see all
	
	return qs.count()


def get_warnings_count(user, from_date, to_date):
	"""Count absence warnings in date range."""
	qs = AbsenceWarning.objects.filter(date__gte=from_date, date__lte=to_date)
	
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			qs = qs.none()
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			qs = qs.filter(employee__department_id=dept_id)
		else:
			qs = qs.none()
	# else: GRH, RH_* see all
	
	return qs.count()


def get_discipline_flags_count(user, from_date, to_date):
	"""Count discipline flags for months in range."""
	qs = DisciplineFlag.objects.filter(month__gte=from_date, month__lte=to_date)
	
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			qs = qs.none()
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			qs = qs.filter(employee__department_id=dept_id)
		else:
			qs = qs.none()
	# else: GRH, RH_* see all
	
	return qs.count()


def get_leaves_counts(user, from_date, to_date):
	"""Return dict with pending/accepted/refused counts."""
	qs = LeaveRequest.objects.filter(start_date__gte=from_date, start_date__lte=to_date)
	
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
			qs = qs.filter(employee_id=emp.id)
		except Employee.DoesNotExist:
			qs = qs.none()
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			qs = qs.filter(employee__department_id=dept_id)
		else:
			qs = qs.none()
	# else: GRH, RH_* see all
	
	return {
		'leaves_pending_count': qs.filter(status=LeaveRequest.Status.PENDING).count(),
		'leaves_accepted_count': qs.filter(status=LeaveRequest.Status.ACCEPTED).count(),
		'leaves_refused_count': qs.filter(status=LeaveRequest.Status.REFUSED).count(),
	}


def get_documents_counts(user, from_date, to_date):
	"""Return dict with created/validated counts."""
	qs = Document.objects.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
	
	if user.role == RoleChoices.EMPLOYEE:
		qs = qs.filter(created_by=user)
	elif user.role == RoleChoices.CHEF:
		dept_id = get_employee_scope(user)
		if dept_id:
			qs = qs.filter(source_department_id=dept_id)
		else:
			qs = qs.none()
	# else: GRH, RH_* see all
	
	return {
		'documents_created_count': qs.count(),
		'documents_validated_count': qs.filter(status=Document.Status.VALIDATED).count(),
	}


def _parse_export_format(request):
	file_format = (request.query_params.get('file_format') or request.query_params.get('format') or 'pdf').lower()
	if file_format not in ('pdf', 'xlsx'):
		return None
	return file_format


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_report(request):
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	employees_qs = _resolve_employee_scope(request.user)
	if employees_qs is None:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	employee_id = request.query_params.get('employee_id')
	if employee_id:
		if request.user.role not in (RoleChoices.CHEF, RoleChoices.GRH):
			return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)
		employees_qs = employees_qs.filter(id=employee_id)

	attendance_qs = Attendance.objects.filter(
		employee__in=employees_qs,
		date__gte=from_date,
		date__lte=to_date
	).select_related('employee')
	attendance_map = {(a.employee_id, a.date): a for a in attendance_qs}
	employees = list(employees_qs.order_by('last_name', 'first_name'))

	rows = []
	current_date = from_date
	while current_date <= to_date:
		for emp in employees:
			attendance = attendance_map.get((emp.id, current_date))
			if attendance and attendance.check_in_time:
				status_label = 'Late' if attendance.check_in_time > EXPORT_LATE_THRESHOLD else 'Present'
			else:
				status_label = 'Absent'

			worked_hours = 0.0
			if attendance and attendance.check_in_time and attendance.check_out_time:
				start_dt = datetime.combine(current_date, attendance.check_in_time)
				end_dt = datetime.combine(current_date, attendance.check_out_time)
				if end_dt < start_dt:
					end_dt = end_dt + timedelta(days=1)
				worked_hours = round((end_dt - start_dt).total_seconds() / 3600, 2)

			rows.append([
				f"{emp.first_name} {emp.last_name}",
				_format_date(current_date),
				_format_time(attendance.check_in_time if attendance else None),
				_format_time(attendance.check_out_time if attendance else None),
				_format_hours(worked_hours),
				status_label,
			])
		current_date += timedelta(days=1)

	headers = ['Employee', 'Date', 'Check-in', 'Check-out', 'Worked Hours', 'Status']
	title = f"Attendance Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"attendance_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		output = _build_excel_report(title, headers, rows)
		return _make_export_response(output, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	output = _build_pdf_report(title, headers, rows)
	return _make_export_response(output, filename, 'application/pdf')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leaves_report(request):
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	user = request.user
	qs = LeaveRequest.objects.all().select_related('employee')
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(employee=emp)
	elif user.role == RoleChoices.CHEF:
		try:
			chef_emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'chef has no employee record'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(employee__department_id=chef_emp.department_id)
	elif user.role != RoleChoices.GRH:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	status_filter = request.query_params.get('status')
	if status_filter:
		qs = qs.filter(status=status_filter.upper())

	type_filter = request.query_params.get('type')
	if type_filter:
		qs = qs.filter(type=type_filter.upper())

	qs = qs.filter(start_date__lte=to_date, end_date__gte=from_date).order_by('start_date')

	rows = []
	for leave in qs:
		days = (leave.end_date - leave.start_date).days + 1
		rows.append([
			f"{leave.employee.first_name} {leave.employee.last_name}",
			leave.type,
			leave.start_date.isoformat(),
			leave.end_date.isoformat(),
			str(days),
			leave.status,
			leave.chef_comment or '-',
		])

	headers = ['Employee', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Chef Comment']
	title = f"Leave History Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"leave_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		output = _build_excel_report(title, headers, rows)
		return _make_export_response(output, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	output = _build_pdf_report(title, headers, rows)
	return _make_export_response(output, filename, 'application/pdf')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_tasks_report(request):
	file_format = _parse_export_format(request)
	if not file_format:
		return Response({'detail': 'invalid format'}, status=status.HTTP_400_BAD_REQUEST)

	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)

	user = request.user
	qs = Task.objects.all().select_related('assigned_by', 'assigned_to')
	if user.role == RoleChoices.EMPLOYEE:
		try:
			emp = Employee.objects.get(email=user.email)
		except Employee.DoesNotExist:
			return Response({'detail': 'employee record not found'}, status=status.HTTP_400_BAD_REQUEST)
		qs = qs.filter(assigned_to=emp)
	elif user.role == RoleChoices.CHEF:
		qs = qs.filter(assigned_by=user)
	elif user.role != RoleChoices.GRH:
		return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)

	qs = qs.filter(created_at__date__gte=from_date, created_at__date__lte=to_date)

	employee_id = request.query_params.get('employee_id')
	if employee_id and user.role in (RoleChoices.CHEF, RoleChoices.GRH):
		qs = qs.filter(assigned_to_id=employee_id)

	status_filter = (request.query_params.get('status') or '').lower()
	if status_filter == 'done':
		qs = qs.filter(status=Task.Status.DONE)
	elif status_filter == 'overdue':
		today = timezone.now().date()
		qs = qs.filter(status=Task.Status.TODO, due_date__lt=today)
	elif status_filter == 'assigned':
		today = timezone.now().date()
		qs = qs.filter(status=Task.Status.TODO).filter(Q(due_date__gte=today) | Q(due_date__isnull=True))

	rows = []
	today = timezone.now().date()
	for task in qs.order_by('-created_at'):
		assigned_by = f"{task.assigned_by.first_name} {task.assigned_by.last_name}".strip() if task.assigned_by else '-'
		assigned_to = f"{task.assigned_to.first_name} {task.assigned_to.last_name}" if task.assigned_to else '-'

		if task.status == Task.Status.DONE:
			status_label = 'Done'
		elif task.due_date and task.due_date < today:
			status_label = 'Overdue'
		else:
			status_label = 'Assigned'

		rows.append([
			task.title,
			task.description or '-',
			assigned_by or '-',
			assigned_to,
			_format_date(task.due_date),
			status_label,
			_format_date(task.completed_at.date() if task.completed_at else None),
		])

	headers = ['Title', 'Description', 'Assigned By', 'Assigned To', 'Due Date', 'Status', 'Completion Date']
	title = f"Task Report ({from_date.isoformat()} to {to_date.isoformat()})"
	filename = f"task_report_{from_date.isoformat()}_{to_date.isoformat()}.{file_format}"

	if file_format == 'xlsx':
		output = _build_excel_report(title, headers, rows)
		return _make_export_response(output, filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	output = _build_pdf_report(title, headers, rows)
	return _make_export_response(output, filename, 'application/pdf')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_summary(request):
	"""
	GET /api/reports/summary/?from=YYYY-MM-DD&to=YYYY-MM-DD
	Returns aggregated counts scoped by user role and department.
	"""
	try:
		from_date, to_date = parse_date_range(request)
	except (ValueError, TypeError):
		return Response({'detail': 'invalid date format (use YYYY-MM-DD)'}, status=status.HTTP_400_BAD_REQUEST)
	
	# Collect all counts
	data = {
		'employees_count': get_employees_count(request.user, from_date, to_date),
		'attendance_days_count': get_attendance_count(request.user, from_date, to_date),
		'absences_detected_count': get_warnings_count(request.user, from_date, to_date),
		'warnings_count': get_warnings_count(request.user, from_date, to_date),
		'discipline_flags_count': get_discipline_flags_count(request.user, from_date, to_date),
	}
	
	# Add leave counts
	data.update(get_leaves_counts(request.user, from_date, to_date))
	
	# Add document counts
	data.update(get_documents_counts(request.user, from_date, to_date))
	
	# Add metadata
	data['from'] = from_date.isoformat()
	data['to'] = to_date.isoformat()
	data['user_role'] = request.user.role
	
	return Response(data)